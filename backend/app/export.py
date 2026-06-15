from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .extraction_quality import assess_document
from .job_summary import SOURCE_LABELS, detect_source, detect_document_subtype, SUBTYPE_LABELS
from .models import JobResult, NormalizedDocument

EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("job_id", "Job ID"),
    ("file_name", "File Name"),
    ("source", "Source"),
    ("document_title", "Document Title"),
    ("document_category", "Document Category"),
    ("data_quality", "Data Quality"),
    ("document_type", "Document Type"),
    ("system_ref_no", "System Ref No"),
    ("line_sl_no", "Sl. No"),
    ("invoice_number", "Invoice Number"),
    ("credit_note_no", "Credit Note No"),
    ("credit_note_date", "Credit Note Date"),
    ("debit_note_no", "Debit Note No"),
    ("debit_note_date", "Debit Note Date"),
    ("invoice_ref_number", "Invoice Reference Number"),
    ("rma_no", "RMA No"),
    ("return_id", "Return ID"),
    ("removal_id", "Removal ID"),
    ("shipment_id", "Shipment ID"),
    ("vret_shipment_id", "VRET Shipment ID"),
    ("due_date", "Due Date"),
    ("payment_method", "Payment Method"),
    ("payment_term", "Payment Term"),
    ("reason", "Reason for Debit Note"),
    ("call_tag_id", "Call Tag ID"),
    ("place_of_supply", "Place of Supply"),
    ("billing_name", "Billing Name"),
    ("billing_address", "Billing Address"),
    ("billing_city", "Billing City"),
    ("billing_state", "Billing State"),
    ("billing_postal_code", "Billing Postal Code"),
    ("billing_country", "Billing Country"),
    ("billing_gstin", "Billing GSTIN"),
    ("billing_state_code", "Billing State Code"),
    ("receiver_billing_name", "Receiver Billing Name"),
    ("receiver_billing_address", "Receiver Billing Address"),
    ("receiver_billing_city", "Receiver Billing City"),
    ("receiver_billing_postal_code", "Receiver Billing Postal Code"),
    ("receiver_billing_gstin", "Receiver Billing GSTIN"),
    ("receiver_billing_state_code", "Receiver Billing State Code"),
    ("receiver_billing_pan", "Receiver Billing PAN"),
    ("shipping_name", "Shipping Name"),
    ("shipping_address", "Shipping Address"),
    ("shipping_city", "Shipping City"),
    ("shipping_state", "Shipping State"),
    ("shipping_postal_code", "Shipping Postal Code"),
    ("shipping_country", "Shipping Country"),
    ("shipping_gstin", "Shipping GSTIN"),
    ("shipping_state_code", "Shipping State Code"),
    ("receiver_shipping_name", "Receiver Shipping Name"),
    ("receiver_shipping_address", "Receiver Shipping Address"),
    ("receiver_shipping_city", "Receiver Shipping City"),
    ("receiver_shipping_postal_code", "Receiver Shipping Postal Code"),
    ("receiver_shipping_gstin", "Receiver Shipping GSTIN"),
    ("receiver_shipping_pan", "Receiver Shipping PAN"),
    ("receiver_shipping_place_of_supply", "Receiver Shipping Place of Supply"),
    ("vendor_name", "Vendor Name"),
    ("vendor_gstin", "Vendor GSTIN"),
    ("vendor_pan", "Vendor PAN"),
    ("customer_name", "Customer Name"),
    ("customer_gstin", "Customer GSTIN"),
    ("customer_pan", "Customer PAN"),
    ("item_description", "Item Description"),
    ("hsn_sac", "HSN/SAC"),
    ("asin_code", "ASIN Code"),
    ("upc_ean", "UPC/EAN"),
    ("purchase_order_no", "Purchase Order No"),
    ("vendor_invoice_no", "Vendor Invoice No"),
    ("vendor_invoice_date", "Vendor Invoice Date"),
    ("unit_code", "Unit/Code"),
    ("quantity", "Quantity"),
    ("rate", "Rate"),
    ("assessable_value", "Assessable Value"),
    ("gst_rate", "GST Rate(%)"),
    ("tax_type", "Tax Type"),
    ("gst_value", "GST Value"),
    ("line_total_amount", "Total Amount"),
]


def _has_field_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return bool(str(value).strip())


def _calculate_data_quality(
    row: dict[str, Any],
    *,
    extraction_status: str | None = None,
) -> str:
    """100% when extraction passed all checks; otherwise score key extracted fields only."""
    if extraction_status == "verified":
        return "100%"

    line_critical = [
        "item_description",
        "hsn_sac",
        "asin_code",
        "line_total_amount",
        "assessable_value",
        "quantity",
        "rate",
        "gst_value",
    ]
    line_optional = [
        "purchase_order_no",
        "vendor_invoice_no",
        "vendor_invoice_date",
        "upc_ean",
    ]
    doc_identifiers = ["credit_note_no", "invoice_number", "debit_note_no"]
    doc_party = ["vendor_gstin", "billing_gstin", "place_of_supply"]

    score = 0
    max_score = len(line_critical) * 2 + len(line_optional) + 3 + len(doc_party)

    for field in line_critical:
        if _has_field_value(row.get(field)):
            score += 2
    for field in line_optional:
        if _has_field_value(row.get(field)):
            score += 1
    if any(_has_field_value(row.get(field)) for field in doc_identifiers):
        score += 3
    for field in doc_party:
        if _has_field_value(row.get(field)):
            score += 1

    pct = round((score / max_score) * 100) if max_score > 0 else 0
    if extraction_status == "failed":
        pct = min(pct, 49)
    elif extraction_status == "needs_review":
        pct = min(pct, 99)
    return f"{pct}%"


def _filename_system_ref(file_name: str) -> str:
    """ETRADE PDFs are often named after the System Ref No (e.g. 30000877790.pdf)."""
    stem = re.sub(r"\s+", "", (file_name or "").rsplit("/", 1)[-1].rsplit("\\", 1)[-1])
    stem = re.sub(r"\.[^.]+$", "", stem, flags=re.I)
    if re.fullmatch(r"\d{8,14}", stem):
        return stem
    return ""


def _document_system_ref_no(header: dict[str, Any], file_name: str, source: str) -> str:
    """Document-level System Ref No (ETRADE header field), not line-item Sl. No."""
    ref = (header.get("system_ref_no") or "").strip()
    if ref:
        return ref
    if source == "etrade":
        return _filename_system_ref(file_name)
    return ""


def _document_number_fields(document: NormalizedDocument, header: dict[str, Any]) -> dict[str, str]:
    """Route invoice / credit note / debit note numbers to the correct export columns."""
    doc_type = document.document_type.value
    invoice_number_out = ""
    credit_note_no = ""
    debit_note_no = ""

    if doc_type == "credit_note":
        credit_note_no = header.get("credit_note_number") or header.get("document_number") or ""
    elif doc_type in ("r4c", "debit_note"):
        debit_note_no = header.get("debit_note_number") or header.get("document_number") or ""
    elif doc_type == "invoice":
        invoice_number_out = header.get("invoice_number") or header.get("document_number") or ""

    return {
        "invoice_number": invoice_number_out,
        "credit_note_no": credit_note_no,
        "credit_note_date": header.get("credit_note_date", ""),
        "debit_note_no": debit_note_no,
        "debit_note_date": header.get("debit_note_date", ""),
    }


def _addr(prefix: str, address: dict[str, Any]) -> dict[str, Any]:
    return {
        f"{prefix}_name": address.get("name", ""),
        f"{prefix}_address": address.get("address", ""),
        f"{prefix}_city": address.get("city", ""),
        f"{prefix}_state": address.get("state", ""),
        f"{prefix}_postal_code": address.get("postal_code", ""),
        f"{prefix}_country": address.get("country", ""),
        f"{prefix}_gstin": address.get("gstin", ""),
        f"{prefix}_state_code": address.get("state_code", ""),
        f"{prefix}_pan": address.get("pan", ""),
        f"{prefix}_place_of_supply": address.get("place_of_supply", ""),
    }


def flatten_document(job_id: str, document: NormalizedDocument) -> list[dict[str, Any]]:
    header = document.header or {}
    audit = document.audit or {}

    source = detect_source(document.vendor.name, document.vendor.gstin, document.vendor.pan)
    reason = header.get("reason") or ""
    file_name = audit.get("file_name", "")
    subtype = detect_document_subtype(document.document_type.value, reason, file_name)
    document_title = header.get("document_heading") or SUBTYPE_LABELS.get(subtype, subtype)
    doc_numbers = _document_number_fields(document, header)

    receiver_pan = header.get("receiver_pan") or document.receiver_billing_address.get("pan", "")
    vendor_pan = header.get("vendor_pan") or document.vendor.pan or ""

    billing_country = document.billing_address.get("country", "") or "India"
    shipping_country = document.shipping_address.get("country", "") or "India"

    base: dict[str, Any] = {
        "job_id": job_id,
        "file_name": file_name,
        "source": SOURCE_LABELS.get(source, source),
        "document_title": document_title,
        "document_category": SUBTYPE_LABELS.get(subtype, subtype),
        "document_type": document.document_type.value,
        "system_ref_no": _document_system_ref_no(header, file_name, source),
        **doc_numbers,
        "invoice_ref_number": header.get("invoice_reference_number", ""),
        "rma_no": header.get("rma_number", ""),
        "return_id": header.get("return_id", ""),
        "removal_id": header.get("removal_id", ""),
        "shipment_id": header.get("shipment_id", ""),
        "vret_shipment_id": header.get("vret_shipment_id", ""),
        "due_date": header.get("due_date", ""),
        "payment_method": header.get("payment_method", ""),
        "payment_term": header.get("payment_terms", ""),
        "reason": header.get("reason", ""),
        "call_tag_id": "",
        "place_of_supply": header.get("place_of_supply", ""),
        "vendor_name": document.vendor.name,
        "vendor_gstin": document.vendor.gstin,
        "vendor_pan": vendor_pan,
        "customer_name": document.customer.name,
        "customer_gstin": document.customer.gstin,
        "customer_pan": document.customer.pan,
    }

    base.update(_addr("billing", document.billing_address))
    base.update(_addr("receiver_billing", document.receiver_billing_address))
    base.update(_addr("shipping", document.shipping_address))
    base.update(_addr("receiver_shipping", document.receiver_shipping_address))

    base["billing_country"] = billing_country
    base["shipping_country"] = shipping_country
    base["receiver_billing_pan"] = receiver_pan or base.get("receiver_billing_pan", "")
    base["receiver_shipping_pan"] = receiver_pan or base.get("receiver_shipping_pan", "")

    extraction_status = assess_document(document).get("extraction_status")

    rows: list[dict[str, Any]] = []
    if not document.line_items:
        row = dict(base)
        row["data_quality"] = _calculate_data_quality(row, extraction_status=extraction_status)
        rows.append(row)
        return rows

    for item in document.line_items:
        row = dict(base)
        row.update(
            {
                "line_sl_no": item.system_ref_no or "",
                "item_description": item.product,
                "hsn_sac": item.hsn,
                "asin_code": item.asin,
                "upc_ean": item.ean or item.sku,
                "purchase_order_no": item.combo,
                "vendor_invoice_no": item.invoice_no,
                "vendor_invoice_date": item.invoice_date,
                "return_id": item.return_id or row.get("return_id", ""),
                "shipment_id": item.shipment_id or row.get("shipment_id", ""),
                "unit_code": item.units,
                "quantity": item.quantity,
                "rate": item.cost_per_unit,
                "assessable_value": item.total_cost,
                "gst_rate": item.tax_rate,
                "tax_type": item.tax_type,
                "gst_value": item.tax_amount,
                "line_total_amount": item.total_amount,
            }
        )
        row["data_quality"] = _calculate_data_quality(row, extraction_status=extraction_status)
        rows.append(row)

    return rows


def flatten_jobs(jobs: list[tuple[str, JobResult]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for job_id, job in jobs:
        if job.document:
            rows.extend(flatten_document(job_id, job.document))
    return rows


def build_xlsx(rows: list[dict[str, Any]], sheet_name: str = "Extracted Data") -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    headers = [label for _, label in EXPORT_COLUMNS]
    keys = [key for key, _ in EXPORT_COLUMNS]
    sheet.append(headers)

    for row in rows:
        sheet.append([row.get(key, "") for key in keys])

    for index, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        column = get_column_letter(index)
        width = min(max(len(label) + 2, 12), 40)
        sheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
